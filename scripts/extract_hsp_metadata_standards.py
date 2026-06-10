#!/usr/bin/env python3
"""Extract local HSP/German metadata standards into YAML and a markdown report."""

from __future__ import annotations

import argparse
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = Path("/Users/mobasuony/Downloads/hsp_normdaten_thesauri (1).xlsx")
DEFAULT_MAPPING_MD = Path("/Users/mobasuony/Downloads/Sources_and_Metadata_data_fields.md")
DEFAULT_TERMS = ROOT / "data/metadata/hsp_normdata_terms.yaml"
DEFAULT_MAPPING = ROOT / "data/metadata/hsp_metadata_field_mapping.yaml"
DEFAULT_REPORT = ROOT / "docs/12_metadata_standards_alignment.md"

VOCABULARIES = {
  "SCRP - Script Types": {
    "code": "SCRP",
    "name": "SCRP script types",
    "expected_term_count": 116,
  },
  "FORM - Fragment & Object Types": {
    "code": "FORM",
    "name": "FORM fragment and object types",
    "expected_term_count": 92,
  },
  "CODC - Codicology": {
    "code": "CODC",
    "name": "CODC codicological concepts",
    "expected_term_count": 73,
  },
  "BNDG - Binding": {
    "code": "BNDG",
    "name": "BNDG binding concepts",
    "expected_term_count": 240,
  },
  "HSP simplified vocabs": {
    "code": "HSP_SIMPLIFIED",
    "name": "HSP simplified controlled vocabularies",
    "expected_term_count": 18,
  },
}


def parse_args() -> argparse.Namespace:
  parser = argparse.ArgumentParser(description="Extract HSP/German normdata metadata references.")
  parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
  parser.add_argument("--mapping", default=str(DEFAULT_MAPPING_MD))
  parser.add_argument("--terms-output", default=str(DEFAULT_TERMS))
  parser.add_argument("--mapping-output", default=str(DEFAULT_MAPPING))
  parser.add_argument("--report", default=str(DEFAULT_REPORT))
  parser.add_argument("--verbose", action="store_true")
  return parser.parse_args()


def now_iso() -> str:
  return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def normalized_sheet_title(title: str) -> str:
  return title.replace("\u2013", "-").replace("\u2014", "-").strip()


def split_allowed_values(value: str | None) -> list[str]:
  if not value or "|" not in value:
    return []
  return [item.strip() for item in value.split("|") if item.strip()]


def code_group_from_notation(notation: str) -> str | None:
  match = re.match(r"^[A-Z]+-([A-Z])\d+", notation)
  if not match:
    return None
  return f"{match.group(1)}-code"


def parse_workbook(path: Path) -> dict[str, Any]:
  if not path.exists():
    raise FileNotFoundError(f"Workbook not found: {path}")

  wb = load_workbook(path, read_only=True, data_only=True)
  vocabularies: list[dict[str, Any]] = []

  for ws in wb.worksheets:
    sheet_key = normalized_sheet_title(ws.title)
    if sheet_key not in VOCABULARIES:
      continue

    config = VOCABULARIES[sheet_key]
    rows = list(ws.iter_rows(values_only=True))
    header_index = None
    for index, row in enumerate(rows):
      if row and row[0] in ("Notation", "term @type"):
        header_index = index
        break
    if header_index is None:
      raise ValueError(f"Could not find header row in sheet {ws.title}")

    terms: list[dict[str, Any]] = []
    for row_index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
      if not row or row[0] is None:
        continue
      first_cell = str(row[0]).strip()
      if not first_cell or first_cell.startswith("\u2500\u2500"):
        continue

      if config["code"] == "HSP_SIMPLIFIED":
        allowed_values_text = str(row[1]).strip() if row[1] is not None else None
        term = {
          "notation": first_cell,
          "label_de": None,
          "label_en": allowed_values_text,
          "code_group": None,
          "norm_uuid": None,
          "allowed_values": split_allowed_values(allowed_values_text),
          "source_sheet": ws.title,
          "source_row": row_index,
          "raw": {
            "term_type": row[0],
            "allowed_values": row[1],
            "notes": row[2],
            "full_thesaurus": row[3],
          },
        }
      else:
        notation = first_cell
        fourth_column = str(row[3]).strip() if row[3] is not None else None
        norm_uuid = fourth_column if fourth_column and fourth_column.startswith("NORM-") else None
        code_group = fourth_column if fourth_column and not fourth_column.startswith("NORM-") else code_group_from_notation(notation)
        term = {
          "notation": notation,
          "label_de": str(row[1]).strip() if row[1] is not None else None,
          "label_en": str(row[2]).strip() if row[2] is not None else None,
          "code_group": code_group,
          "norm_uuid": norm_uuid,
          "allowed_values": [],
          "source_sheet": ws.title,
          "source_row": row_index,
          "raw": {
            "notation": row[0],
            "label_de": row[1],
            "label_en": row[2],
            "code_group_or_uuid": row[3],
          },
        }
      terms.append(term)

    vocabularies.append(
      {
        "code": config["code"],
        "name": config["name"],
        "source_sheet": ws.title,
        "expected_term_count": config["expected_term_count"],
        "actual_term_count": len(terms),
        "terms": terms,
      }
    )

  return {
    "generated_at": now_iso(),
    "source_workbook": str(path),
    "vocabularies": vocabularies,
  }


def parse_md_table_row(line: str) -> list[str]:
  return [cell.strip() for cell in line.strip().strip("|").split("|")]


def parse_mapping_markdown(path: Path) -> dict[str, Any]:
  if not path.exists():
    raise FileNotFoundError(f"Mapping markdown not found: {path}")

  lines = path.read_text(encoding="utf-8").splitlines()
  tables: list[dict[str, Any]] = []
  current_table: str | None = None
  current_section: str | None = None
  pending_table: list[str] = []

  def flush_pending() -> None:
    nonlocal pending_table
    if not pending_table:
      return
    header = parse_md_table_row(pending_table[0])
    rows = []
    for raw_line in pending_table[2:]:
      cells = parse_md_table_row(raw_line)
      if len(cells) != len(header):
        continue
      row = dict(zip(header, cells, strict=True))
      rows.append(row)
    if rows:
      table_name = current_table or "unassigned"
      for row in rows:
        row["source_table"] = table_name
        row["source_section"] = current_section
      tables.append({"table": table_name, "section": current_section, "rows": rows})
    pending_table = []

  for line in lines:
    table_match = re.match(r"^## Table: `([^`]+)`", line)
    if table_match:
      flush_pending()
      current_table = table_match.group(1)
      current_section = None
      continue

    if line.startswith("## Fields with no natural home"):
      flush_pending()
      current_table = "fragment_extensions"
      current_section = "Fields with no natural home"
      continue

    section_match = re.match(r"^###\s+(.+)$", line)
    if section_match:
      flush_pending()
      current_section = section_match.group(1)
      continue

    if line.startswith("|"):
      pending_table.append(line)
    else:
      flush_pending()

  flush_pending()

  fields: list[dict[str, Any]] = []
  for table in tables:
    for row in table["rows"]:
      fields.append(row)

  return {
    "generated_at": now_iso(),
    "source_markdown": str(path),
    "table_count": len(tables),
    "field_count": len(fields),
    "tables": tables,
  }


def write_yaml(path: Path, payload: dict[str, Any]) -> None:
  path.parent.mkdir(parents=True, exist_ok=True)
  path.write_text(yaml.safe_dump(payload, sort_keys=False, allow_unicode=True), encoding="utf-8")


def write_report(path: Path, terms: dict[str, Any], mapping: dict[str, Any]) -> None:
  lines = [
    "# Fragment Autocomplete - Metadata Standards Alignment",
    "",
    "## Purpose",
    "",
    "This report records the HSP/German cataloguing normdata alignment inputs used by the local pipeline. It preserves the existing UUID/PostGIS schema while adding controlled-vocabulary support and selected normalized catalogue fields.",
    "",
    "## Source Inputs",
    "",
    f"- Workbook: `{terms['source_workbook']}`",
    f"- Field mapping: `{mapping['source_markdown']}`",
    "",
    "## Controlled Vocabularies",
    "",
    "| Vocabulary | Sheet | Expected terms | Extracted terms |",
    "| --- | --- | ---: | ---: |",
  ]
  for vocab in terms["vocabularies"]:
    lines.append(
      f"| `{vocab['code']}` | {vocab['source_sheet']} | {vocab['expected_term_count']} | {vocab['actual_term_count']} |"
    )
  lines.extend(
    [
      "",
      "## Field Mapping Summary",
      "",
      f"- Parsed mapping tables: `{mapping['table_count']}`",
      f"- Parsed mapping fields: `{mapping['field_count']}`",
      "- The mapping is treated as a standards reference, not as a direct replacement for the implemented database schema.",
      "- Existing source metadata remains in `raw_metadata`; normalized fields are populated only when values are detectable or explicitly reviewed.",
      "",
      "## Pipeline Implications",
      "",
      "- IIIF ingestion and local sample registration should populate HSP-aligned fields opportunistically.",
      "- Rights and access flags remain conservative until explicit review.",
      "- Controlled notation assignments should reference imported `controlled_term` rows rather than free text alone.",
      "- Segmentation, artificial-fragment generation, retrieval, and reconstruction can use normalized metadata as context, not as proof of historical reconstruction.",
      "",
      "## Next Step",
      "",
      "Import these vocabularies into PostgreSQL/PostGIS and validate the metadata alignment layer before expanding dataset registration.",
      "",
    ]
  )
  path.parent.mkdir(parents=True, exist_ok=True)
  path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
  args = parse_args()
  workbook = Path(args.workbook)
  mapping_md = Path(args.mapping)
  terms_output = Path(args.terms_output)
  mapping_output = Path(args.mapping_output)
  report_path = Path(args.report)

  terms = parse_workbook(workbook)
  mapping = parse_mapping_markdown(mapping_md)
  write_yaml(terms_output, terms)
  write_yaml(mapping_output, mapping)
  write_report(report_path, terms, mapping)

  if args.verbose:
    print(f"Wrote {terms_output}")
    print(f"Wrote {mapping_output}")
    print(f"Wrote {report_path}")
    for vocab in terms["vocabularies"]:
      print(f"{vocab['code']}: {vocab['actual_term_count']} terms")
    print(f"Mapping fields: {mapping['field_count']}")
  return 0


if __name__ == "__main__":
  raise SystemExit(main())
